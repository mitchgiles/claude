"""
Exports gluten-free price comparison results to a formatted Excel workbook.
"""

from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ── Palette ──────────────────────────────────────────────────────────────────
GREEN_DARK = "1B5E20"
GREEN_MID = "2E7D32"
GREEN_LIGHT = "C8E6C9"
GREEN_ROW1 = "E8F5E9"
GREEN_ROW2 = "F1F8E9"
BLUE_LINK = "1565C0"
RED_LOSS = "C62828"
GREY_BORDER = "CCCCCC"
WHITE = "FFFFFF"

# ── Reusable style helpers ────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color=GREY_BORDER),
    right=Side(style="thin", color=GREY_BORDER),
    top=Side(style="thin", color=GREY_BORDER),
    bottom=Side(style="thin", color=GREY_BORDER),
)

HEADER_FONT = Font(bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=GREEN_MID, end_color=GREEN_MID, fill_type="solid")
SUMMARY_FILL = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
CURRENCY_FMT = '"$"#,##0.00'

GF_STATUS_LABELS = {
    "naturally_gf": "Naturally GF",
    "labeled_gf": "Labeled GF",
    "has_gf_alternative": "GF Alternative",
}

MAIN_HEADERS = [
    "Item (as on Receipt)",
    "Category",
    "GF Status",
    "Paid (CAD)",
    "Lowest CA Price",
    "Best Store (Canada)",
    "Best Matching Product",
    "Savings (CAD)",
    "Product Link",
]

COMP_HEADERS = ["Receipt Item", "Store", "Price (CAD)", "Product / Size"]


def _set_header_row(ws, row: int, headers: list[str]):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 26


def _apply_border_and_fill(cell, fill, number_format=None, align=None):
    cell.border = THIN_BORDER
    cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if align:
        cell.alignment = align


def build_main_sheet(ws, receipt_data: dict, results: list[dict]):
    store = receipt_data.get("store_name", "Unknown Store")
    date = receipt_data.get("receipt_date", "Unknown Date")
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Title ─────────────────────────────────────────────────────────────────
    num_cols = len(MAIN_HEADERS)
    last_col = get_column_letter(num_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "Gluten-Free Price Finder — Canada"
    ws["A1"].font = Font(bold=True, size=15, color=GREEN_DARK)
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Receipt: {store}  |  Date: {date}  |  Generated: {generated}"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 18

    # ── Headers ───────────────────────────────────────────────────────────────
    HEADER_ROW = 4
    _set_header_row(ws, HEADER_ROW, MAIN_HEADERS)

    # ── Data ──────────────────────────────────────────────────────────────────
    total_paid = 0.0
    total_lowest = 0.0
    total_savings = 0.0

    for i, item in enumerate(results):
        row = HEADER_ROW + 1 + i
        fill = PatternFill(
            start_color=GREEN_ROW1 if i % 2 == 0 else GREEN_ROW2,
            end_color=GREEN_ROW1 if i % 2 == 0 else GREEN_ROW2,
            fill_type="solid",
        )

        price_paid = item.get("price_paid") or 0.0
        lowest = item.get("lowest_price_cad") or price_paid
        savings = item.get("savings_cad", 0.0) or 0.0

        total_paid += price_paid
        total_lowest += lowest
        total_savings += savings

        status_raw = item.get("gluten_free_status", "")
        status_label = GF_STATUS_LABELS.get(status_raw, status_raw.replace("_", " ").title())

        url = item.get("lowest_price_url") or ""

        row_values = [
            item.get("receipt_name", ""),
            item.get("category", "").title(),
            status_label,
            price_paid,
            lowest,
            item.get("lowest_price_store", "N/A"),
            item.get("lowest_price_product", "N/A"),
            savings,
            url,
        ]

        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [1, 7, 9]))

            # Currency columns
            if col in (4, 5, 8):
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Colour savings
            if col == 8:
                if savings > 0:
                    cell.font = Font(bold=True, color=GREEN_MID)
                elif savings < 0:
                    cell.font = Font(bold=True, color=RED_LOSS)

            # Hyperlink
            if col == 9 and url.startswith("http"):
                cell.hyperlink = url
                cell.value = "View Product"
                cell.font = Font(color=BLUE_LINK, underline="single")

        ws.row_dimensions[row].height = 28

    # ── Summary row ───────────────────────────────────────────────────────────
    summary_row = HEADER_ROW + len(results) + 2
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    label_cell = ws[f"A{summary_row}"]
    label_cell.value = "TOTALS"
    label_cell.font = Font(bold=True, size=11, color=GREEN_DARK)
    label_cell.alignment = CENTER
    label_cell.fill = SUMMARY_FILL

    for col, val in ((4, total_paid), (5, total_lowest), (8, total_savings)):
        cell = ws.cell(row=summary_row, column=col, value=val)
        cell.font = Font(bold=True, color=GREEN_DARK)
        cell.number_format = CURRENCY_FMT
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    ws.row_dimensions[summary_row].height = 24

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [32, 14, 18, 14, 16, 26, 38, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    return total_paid, total_lowest, total_savings


def build_comparison_sheet(ws, results: list[dict]):
    ws["A1"] = "Store-by-Store Price Comparisons"
    ws["A1"].font = Font(bold=True, size=13, color=GREEN_DARK)
    ws.row_dimensions[1].height = 24

    _set_header_row(ws, 3, COMP_HEADERS)

    row = 4
    for item in results:
        comparisons = item.get("comparison_prices") or []
        for comp in comparisons:
            fill = PatternFill(
                start_color=GREEN_ROW1 if row % 2 == 0 else GREEN_ROW2,
                end_color=GREEN_ROW1 if row % 2 == 0 else GREEN_ROW2,
                fill_type="solid",
            )
            values = [
                item.get("receipt_name", ""),
                comp.get("store", ""),
                comp.get("price_cad"),
                comp.get("product", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = THIN_BORDER
                cell.fill = fill
                if col == 3:
                    cell.number_format = CURRENCY_FMT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            ws.row_dimensions[row].height = 22
            row += 1

    for i, w in enumerate([32, 28, 14, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=4, column=1)


def build_notes_sheet(ws, results: list[dict]):
    ws["A1"] = "Search Notes & Details"
    ws["A1"].font = Font(bold=True, size=13, color=GREEN_DARK)
    ws.row_dimensions[1].height = 24

    _set_header_row(ws, 3, ["Receipt Item", "Search Query Used", "Search Notes"])

    row = 4
    for item in results:
        fill = PatternFill(
            start_color=GREEN_ROW1 if row % 2 == 0 else GREEN_ROW2,
            end_color=GREEN_ROW1 if row % 2 == 0 else GREEN_ROW2,
            fill_type="solid",
        )
        for col, val in enumerate(
            [
                item.get("receipt_name", ""),
                item.get("search_query", ""),
                item.get("search_notes", ""),
            ],
            1,
        ):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    for i, w in enumerate([32, 40, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=4, column=1)


def export_to_excel(receipt_data: dict, results: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    ws_main = wb.active
    ws_main.title = "Price Comparison"

    ws_comp = wb.create_sheet("Store Comparisons")
    ws_notes = wb.create_sheet("Search Notes")

    total_paid, total_lowest, total_savings = build_main_sheet(ws_main, receipt_data, results)
    build_comparison_sheet(ws_comp, results)
    build_notes_sheet(ws_notes, results)

    wb.save(output_path)
    return total_paid, total_lowest, total_savings
